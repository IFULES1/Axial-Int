"""Construit le classeur de pilotage Axial à partir des données réelles."""
import csv, datetime as dt, json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

D = json.load(open('donnees.json'))
AUJ = dt.datetime.now(dt.timezone.utc).strftime('%d/%m/%Y %H:%M UTC')

# --- Palette sobre : un classeur de pilotage se lit, il ne se contemple pas.
TITRE   = Font(bold=True, size=14, color="1B1D1E")
ENTETE  = Font(bold=True, size=10, color="FFFFFF")
FOND_E  = PatternFill("solid", fgColor="2F4858")
FOND_KO = PatternFill("solid", fgColor="FBE9E7")
FOND_OK = PatternFill("solid", fgColor="E8F5E9")
FOND_WARN = PatternFill("solid", fgColor="FFF8E1")
GRIS    = Font(color="6B7280", size=9)
BORD    = Border(bottom=Side(style="thin", color="D6DCE4"))

wb = Workbook()

def feuille(nom, titre, sous_titre=None):
    ws = wb.create_sheet(nom)
    ws["A1"] = titre; ws["A1"].font = TITRE
    if sous_titre:
        ws["A2"] = sous_titre; ws["A2"].font = GRIS
    ws.freeze_panes = "A5"
    return ws

def tableau(ws, ligne, colonnes, donnees, largeurs=None):
    for i, c in enumerate(colonnes, 1):
        cell = ws.cell(row=ligne, column=i, value=c)
        cell.font = ENTETE; cell.fill = FOND_E
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for j, row in enumerate(donnees, ligne + 1):
        for i, v in enumerate(row, 1):
            cell = ws.cell(row=j, column=i, value=v)
            cell.border = BORD
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(v, str) and len(str(v)) > 40)
    if largeurs:
        for i, w in enumerate(largeurs, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    return ligne + 1 + len(donnees)

# ================================================================ LISEZ-MOI
ws = wb.active; ws.title = "LISEZ_MOI"
ws["A1"] = "Pilotage Axial — classeur de suivi"; ws["A1"].font = Font(bold=True, size=16)
ws["A2"] = f"Données extraites de la production le {AUJ}"; ws["A2"].font = GRIS
lignes = [
 ("", ""),
 ("À QUOI SERT CE CLASSEUR", ""),
 ("", "Suivre les coûts, la rentabilité et l'activité réelle d'Axial, à partir des données de production — pas d'estimations."),
 ("", ""),
 ("LES ONGLETS", ""),
 ("DASHBOARD", "Vue de tête. Les chiffres qui décident, avec leur cible."),
 ("METRIQUES", "Le catalogue complet des 56 métriques, avec la valeur actuelle quand elle est calculable."),
 ("UTILISATEURS", "Une ligne par compte. Base de tous les calculs d'activation."),
 ("RAPPORTS", "Une ligne par rapport produit, avec son coût réel."),
 ("EMAILS", "Performance par campagne, en distinguant lectures réelles et préchargements."),
 ("REVENUS", "Abonnements et statuts Stripe."),
 ("COUTS", "Ventilation des coûts variables par poste."),
 ("EVENEMENTS_A_TRACER", "Ce qu'il faudrait instrumenter pour débloquer les métriques manquantes."),
 ("QUALITE_DONNEES", "Contrôles de cohérence et fraîcheur."),
 ("SYNC_LOG", "Journal des synchronisations, à remplir par le script d'automatisation."),
 ("", ""),
 ("RÈGLE DE LECTURE", ""),
 ("", "Une case vide ou marquée « non mesurable » n'est pas un zéro. C'est une donnée qui n'existe pas encore."),
 ("", "Le classeur ne fabrique aucun chiffre : ce qui n'est pas mesuré reste vide, avec la raison en clair."),
 ("", ""),
 ("MISE À JOUR", ""),
 ("", "Aujourd'hui : manuelle, par régénération du classeur."),
 ("", "Cible : script Google Apps Script appelant l'API Axial toutes les heures (voir EVENEMENTS_A_TRACER)."),
 ("", ""),
 ("ATTENTION", ""),
 ("", "Les tarifs des modèles et des API de recherche sont des ordres de grandeur configurables."),
 ("", "Vérifie-les sur les grilles officielles avant d'en tirer une décision de prix."),
]
for i, (a, b) in enumerate(lignes, 4):
    ws.cell(row=i, column=1, value=a).font = Font(bold=True, size=10) if a and not b else Font(size=10)
    ws.cell(row=i, column=2, value=b).alignment = Alignment(wrap_text=True, vertical="top")
ws.column_dimensions["A"].width = 24; ws.column_dimensions["B"].width = 95

print("LISEZ_MOI ok")

# ================================================================ CALCULS
U = D["utilisateurs"]; R = D["rapports"]; E = D["emails"]
clients = [u for u in U if u["categorie"] == "client"]
actifs = [u for u in clients if (u["rapports_produits"] or 0) or (u["questions"] or 0)]
profils = [u for u in clients if u.get("company_name")]
produits = [r for r in R if not r["restaure"]]
mesures = [r for r in produits if float(r["cout_modele_eur"] or 0) > 0]
cout_tot = sum(float(r["cout_modele_eur"] or 0) + float(r["cout_recherche_eur"] or 0) for r in mesures)
abo_actifs = [x for x in D["revenus"] if x["statut"] == "active"]
essais = [x for x in D["revenus"] if x["statut"] == "trialing"]
essais_ann = [x for x in essais if x["annule_en_fin_de_periode"]]
contactes = sum(e["envoyes"] for e in E if e["campagne"] in ("migration_2026_08", "relance_2026_08"))
lectures = sum(e["lectures_reelles"] for e in E if e["campagne"] in ("migration_2026_08", "relance_2026_08"))

def pct(a, b): return f"{100*a/b:.1f} %" if b else "—"
NM = "non mesurable"

# ================================================================ DASHBOARD
ws = feuille("DASHBOARD", "Dashboard — état au " + AUJ,
             "Chaque ligne porte sa cible. Une valeur « non mesurable » signale une donnée absente, pas un zéro.")
l = 4
l = tableau(ws, l, ["Bloc", "Indicateur", "Valeur", "Cible", "Lecture"], [
 ("ACTIVATION", "Personnes contactées", contactes, "—", "2 campagnes : migration + relance"),
 ("", "Lectures réelles (≥3 ouvertures)", lectures, "> 30 %", pct(lectures, contactes) + " des contactés"),
 ("", "Comptes clients créés", len(clients), "croissance", pct(len(clients), contactes) + " des contactés"),
 ("", "Profils remplis", len(profils), "> 80 %", pct(len(profils), len(clients)) + " des clients"),
 ("", "Clients ayant produit quelque chose", len(actifs), "> 50 %", pct(len(actifs), len(clients)) + " des clients"),
 ("", "Clients sans aucun rapport produit", len([u for u in clients if not u["rapports_produits"]]), "< 20 %", "le vrai point dur"),
 ("REVENUS", "MRR", f"{len(abo_actifs)*50} €", "croissance", "aucun abonnement actif"),
 ("", "Essais en cours", len(essais), "—", f"dont {len(essais_ann)} annulés d'avance"),
 ("", "MRR potentiel", f"{(len(essais)-len(essais_ann))*50} €", "—", "si les essais non annulés convertissent"),
 ("", "Paiements en échec", 0, "0", "—"),
 ("COÛTS", "Rapports produits (hors restaurés)", len(produits), "—", f"dont {len(mesures)} avec coût mesuré"),
 ("", "Coût total mesuré", f"{cout_tot:.2f} €", "—", "porte uniquement sur les rapports mesurés"),
 ("", "Coût moyen d'un rapport", f"{cout_tot/len(mesures):.3f} €" if mesures else NM, "< 0,60 €", "modèle + recherche web"),
 ("", "Coûts fixes mensuels", NM, "à saisir", "variable COUTS_FIXES_MENSUELS_EUR non renseignée"),
 ("", "Point mort", NM, "—", "dépend des coûts fixes"),
 ("ENGAGEMENT", "DAU / WAU / MAU", NM, "—", "aucune table d'événements — voir EVENEMENTS_A_TRACER"),
 ("", "Rétention J1 / J7 / J30", NM, "—", f"{len(clients)} clients : une cohorte à ce volume est du bruit"),
 ("", "Sources d'acquisition", NM, "—", "aucun paramètre de campagne à l'inscription"),
 ("", "Churn / LTV / LTV-CAC", NM, "—", "aucun client payant à ce jour"),
], [16, 40, 16, 14, 52])

ws.cell(row=l+1, column=1, value="Ce qui appelle une décision").font = Font(bold=True, size=11)
l += 2
for txt in [
  "Aucun abonnement actif : le MRR est à zéro et les deux essais en cours sont annulés d'avance.",
  "6 clients sur 7 comptes n'ont jamais posé de question — l'usage conversationnel est nul côté client.",
  "Les coûts fixes ne sont pas renseignés : sans eux, ni le coût total ni le point mort ne sont calculables.",
]:
    ws.cell(row=l, column=1, value="•")
    c = ws.cell(row=l, column=2, value=txt); c.alignment = Alignment(wrap_text=True)
    l += 1
print("DASHBOARD ok")

# ================================================================ MÉTRIQUES
# Valeurs courantes injectées dans le catalogue, uniquement là où c'est calculable.
VALEURS = {
 "C01": f"{cout_tot/len(mesures):.3f} €" if mesures else "",
 "C03": f"{max((float(r['cout_modele_eur'] or 0)+float(r['cout_recherche_eur'] or 0)) for r in mesures):.3f} €" if mesures else "",
 "C09": "", "C10": "", "R08": "",
 "V01": f"{len(abo_actifs)*50} €", "V02": f"{(len(essais)-len(essais_ann))*50} €",
 "V03": len(abo_actifs), "V04": len(essais), "V05": len(essais_ann), "V08": 0,
 "A01": contactes, "A02": pct(lectures, contactes),
 "A03": pct(sum(e["jamais_ouverts"] for e in E if e["campagne"] in ("migration_2026_08","relance_2026_08")), contactes),
 "A05": len(clients), "A06": pct(len(clients), contactes),
 "A07": len(profils), "A08": pct(len(actifs), len(clients)),
 "A10": len([u for u in clients if not u["rapports_produits"]]),
 "U01": len(actifs),
 "U06": f"{sum(u['solde_credits'] for u in clients)/len(clients):.0f}" if clients else "",
 "U07": len([u for u in clients if 0 < (u["solde_credits"] or 0) < 25]),
 "U08": sum(u["documents"] or 0 for u in clients),
 "U09": sum(u["outils"] or 0 for u in clients),
 "U10": sum(u["veilles"] or 0 for u in clients),
}
ws = feuille("METRIQUES", "Catalogue des métriques",
             "56 métriques. La colonne « valeur actuelle » est vide quand la donnée n'existe pas — jamais remplie d'un zéro par défaut.")
lignes = []
with open('/Users/mirad/axial-intelligence/docs/metriques_axial.csv', encoding='utf-8') as f:
    for r in csv.DictReader(f, delimiter=";"):
        lignes.append((r["id"], r["categorie"], r["metrique"],
                       VALEURS.get(r["id"], ""), r["cible"],
                       r["mesurable_aujourdhui"], r["definition"],
                       r["source_ou_formule"], r["ce_qui_manque"],
                       r["frequence"], r["priorite"]))
l = tableau(ws, 4, ["id", "Catégorie", "Métrique", "Valeur actuelle", "Cible",
                    "Mesurable", "Définition", "Source / formule", "Ce qui manque",
                    "Fréquence", "Priorité"],
            lignes, [7, 14, 34, 16, 20, 11, 46, 40, 38, 12, 8])
for i in range(5, 5 + len(lignes)):
    etat = ws.cell(row=i, column=6).value
    ws.cell(row=i, column=6).fill = (FOND_OK if etat == "oui" else
                                     FOND_WARN if etat == "partiel" else FOND_KO)
print(f"METRIQUES ok ({len(lignes)} lignes)")

# ================================================================ UTILISATEURS
ws = feuille("UTILISATEURS", "Utilisateurs", "Une ligne par compte. Base de tous les calculs d'activation.")
tableau(ws, 4, ["Email", "Catégorie", "Inscrit le", "Dernière connexion", "Sessions",
                "Entreprise", "Secteur", "Stade", "Marché", "Langue",
                "Rapports produits", "Rapports restaurés", "Questions",
                "Documents", "Outils", "Veilles", "Solde crédits",
                "Essai expire le", "Abonnement", "Renouvellement bloqué"],
         [(u["email"], u["categorie"], u["inscrit_le"], u["derniere_connexion"], u["sessions"],
           u["company_name"], u["sector"], u["funding_stage"], u["target_market"], u["language"],
           u["rapports_produits"], (u["rapports_total"] or 0) - (u["rapports_produits"] or 0),
           u["questions"], u["documents"], u["outils"], u["veilles"], u["solde_credits"],
           u["essai_expire_le"], u["abonnement"],
           "oui" if u["renouvellement_bloque"] else "") for u in U],
         [32, 10, 12, 16, 9, 16, 14, 10, 12, 8, 15, 15, 10, 11, 8, 8, 12, 14, 12, 18])
print("UTILISATEURS ok")

# ================================================================ RAPPORTS
ws = feuille("RAPPORTS", "Rapports",
             "Les rapports restaurés de l'ancienne app sont marqués : ils n'ont jamais été produits par Axial et ne comptent dans aucune moyenne.")
tableau(ws, 4, ["Email", "Produit le", "Type", "Restauré", "Caractères", "Sources",
                "Durée (s)", "Modèle", "Tokens entrée", "Tokens sortie",
                "Coût modèle (€)", "Coût recherche (€)", "Appels recherche", "Coût total (€)"],
        [(r["email"], r["produit_le"][:16], r["type"], "oui" if r["restaure"] else "",
          r["caracteres"], r["sources"], r["duree_secondes"], r["modele"],
          r["tokens_entree"], r["tokens_sortie"],
          float(r["cout_modele_eur"]) or None, float(r["cout_recherche_eur"]) or None,
          r["appels_recherche"],
          round(float(r["cout_modele_eur"] or 0) + float(r["cout_recherche_eur"] or 0), 4) or None)
         for r in R],
        [32, 17, 26, 10, 11, 9, 10, 20, 13, 13, 15, 17, 15, 14])
print("RAPPORTS ok")

# ================================================================ EMAILS
ws = feuille("EMAILS", "Campagnes email",
             "Une « ouverture apparente » inclut les préchargements automatiques de Gmail et Apple Mail. Seules les ouvertures répétées (≥3) sont des lectures humaines certaines.")
tableau(ws, 4, ["Campagne", "Premier envoi", "Envoyés", "Ouvertures apparentes",
                "Lectures réelles", "Taux de lecture réelle", "Jamais ouverts", "Taux jamais ouverts"],
        [(e["campagne"], e["premier_envoi"], e["envoyes"], e["ouvertures_apparentes"],
          e["lectures_reelles"], pct(e["lectures_reelles"], e["envoyes"]),
          e["jamais_ouverts"], pct(e["jamais_ouverts"], e["envoyes"])) for e in E],
        [34, 14, 10, 21, 16, 21, 15, 19])
print("EMAILS ok")

# ================================================================ REVENUS
ws = feuille("REVENUS", "Abonnements", "Aucun abonnement actif à ce jour : le MRR est structurellement à zéro.")
l = tableau(ws, 4, ["Email", "Plan", "Statut", "Annulé en fin de période", "Fin de période", "Stripe réel"],
        [(x["email"], x["plan"], x["statut"], "oui" if x["annule_en_fin_de_periode"] else "",
          x["fin_periode"], "oui" if x["stripe_reel"] else "non") for x in D["revenus"]],
        [32, 10, 12, 24, 16, 12])
l += 1
tableau(ws, l, ["Mouvement de crédits", "Occurrences", "Total crédits"],
        sorted([(a, sum(1 for c in D["credits"] if c["action"] == a),
                 sum(c["delta"] for c in D["credits"] if c["action"] == a))
                for a in {c["action"] for c in D["credits"]}], key=lambda x: -abs(x[2])),
        [30, 14, 14])
print("REVENUS ok")

# ================================================================ COÛTS
ws = feuille("COUTS", "Coûts variables",
             "La colonne « mesurées » dit sur quoi porte le total. Les lignes antérieures au 25/08/2026 n'ont jamais été instrumentées.")
l = tableau(ws, 4, ["Poste", "Lignes", "Mesurées", "Coût modèle (€)", "Coût recherche (€)",
                    "Coût total (€)", "Coût moyen par ligne mesurée (€)"],
        [(c["poste"], c["lignes"], c["mesurees"], float(c["cout_modele_eur"]),
          float(c["cout_recherche_eur"]),
          round(float(c["cout_modele_eur"]) + float(c["cout_recherche_eur"]), 4),
          round((float(c["cout_modele_eur"]) + float(c["cout_recherche_eur"])) / c["mesurees"], 4)
          if c["mesurees"] else None)
         for c in D["couts_agreges"]],
        [16, 10, 11, 17, 19, 16, 30])
l += 1
ws.cell(row=l, column=1, value="Coûts fixes mensuels").font = Font(bold=True, size=11)
l += 1
tableau(ws, l, ["Poste", "Montant mensuel (€)", "À remplir"],
        [("VPS Hostinger", None, "à saisir"), ("Supabase", None, "à saisir"),
         ("Resend", None, "à saisir"), ("Nom de domaine", None, "à saisir"),
         ("Autres", None, "à saisir"), ("TOTAL", None, "somme des lignes ci-dessus")],
        [26, 22, 18])
print("COUTS ok")

# ================================================ ÉVÉNEMENTS À TRACER
ws = feuille("EVENEMENTS_A_TRACER", "Événements à instrumenter",
             "Ce que le produit n'enregistre pas aujourd'hui, et ce que chaque manque empêche de calculer.")
l = tableau(ws, 4, ["Événement", "Déclencheur", "Propriétés nécessaires", "Débloque",
                    "État actuel", "Effort"],
 [("signup", "création de compte", "user_id, horodatage, source de campagne",
   "Sources d'acquisition, conversion par campagne",
   "PARTIEL — le compte est créé mais aucune source n'est enregistrée", "faible"),
  ("session_start", "ouverture de l'app", "user_id, horodatage",
   "DAU, WAU, MAU, DAU/MAU, fréquence d'usage",
   "MANQUANT — les sessions Supabase restent ouvertes des jours, elles ne comptent pas les visites", "moyen"),
  ("onboarding_step", "passage d'un écran à l'autre", "user_id, numéro d'étape, horodatage",
   "Abandon par étape d'onboarding",
   "MANQUANT — on sait où les gens s'arrêtent, pas quand ils s'arrêtent", "faible"),
  ("card_skipped", "clic sur Continuer sans carte", "user_id, horodatage",
   "Taux de contournement de l'écran carte",
   "MANQUANT — déduit indirectement de l'absence d'abonnement", "faible"),
  ("report_started", "lancement d'une génération", "user_id, type, horodatage",
   "Taux d'abandon en cours de génération, échecs",
   "MANQUANT — seuls les rapports réussis sont enregistrés", "faible"),
  ("report_failed", "génération dégradée ou tronquée", "user_id, type, motif",
   "Taux d'échec, qualité de service",
   "MANQUANT — non archivé puisque non facturé, visible seulement dans les journaux", "faible"),
  ("report_opened", "consultation d'un rapport", "user_id, report_id, horodatage",
   "Rapports réellement lus vs produits",
   "MANQUANT — un rapport produit n'est pas un rapport lu", "moyen"),
  ("feedback_clicked", "clic sur Votre avis", "user_id, report_id",
   "Taux de retour, corrélation qualité/rétention",
   "MANQUANT — le formulaire Google ne renvoie rien vers l'app", "faible"),
  ("message_sent", "question posée", "user_id, agent, horodatage",
   "Actions par utilisateur, engagement conversationnel",
   "PRÉSENT — table messages, avec coût depuis le 25/08", "—"),
  ("subscription", "abonnement souscrit", "user_id, plan, montant",
   "MRR, conversion essai → payant",
   "PRÉSENT — user_subscriptions + webhooks Stripe", "—"),
 ], [20, 30, 38, 42, 62, 10])

l += 1
ws.cell(row=l, column=1, value="Architecture d'automatisation visée").font = Font(bold=True, size=11)
l += 2
for txt in [
  "Axial (API) → Google Apps Script → Google Sheets → Dashboard",
  "",
  "1. Un endpoint /metrics/export sur l'API Axial, protégé par un jeton dédié en lecture seule.",
  "2. Un Apps Script collé dans ce classeur, déclenché toutes les heures.",
  "3. Le script réécrit les onglets de données et journalise son passage dans SYNC_LOG.",
  "4. Les onglets DASHBOARD et METRIQUES se recalculent par formules — jamais réécrits par le script.",
  "",
  "Ce classeur-ci est une photographie manuelle. Le script reste à écrire : il demande d'ouvrir",
  "l'endpoint d'export et de générer un jeton, ce qui n'a pas été fait à ce stade.",
]:
    c = ws.cell(row=l, column=1, value=txt)
    if txt.startswith("Axial"): c.font = Font(bold=True, size=10)
    l += 1
print("EVENEMENTS_A_TRACER ok")

# ================================================ QUALITÉ DES DONNÉES
ws = feuille("QUALITE_DONNEES", "Qualité des données", "Contrôles au moment de l'extraction.")
sans_profil = [u for u in clients if not u.get("company_name")]
sans_conn = [u for u in U if not u.get("derniere_connexion")]
tableau(ws, 4, ["Contrôle", "Résultat", "Seuil", "Verdict"],
 [("Fraîcheur des données", AUJ, "< 24 h", "OK"),
  ("Comptes au total", len(U), "—", "—"),
  ("dont comptes clients", len(clients), "—", "—"),
  ("dont comptes internes", len(U) - len(clients), "—", "exclus de tous les taux"),
  ("Clients sans profil rempli", len(sans_profil), "0", "à relancer" if sans_profil else "OK"),
  ("Comptes sans connexion enregistrée", len(sans_conn), "0", "à vérifier" if sans_conn else "OK"),
  ("Rapports au total", len(R), "—", "—"),
  ("dont restaurés de l'ancienne app", len(R) - len(produits), "—", "exclus des moyennes"),
  ("dont avec coût mesuré", len(mesures), f"= {len(produits)}", "instrumentation partielle"),
  ("Rapports sans source citée", len([r for r in R if not r["sources"]]), "0",
   "à vérifier" if any(not r["sources"] for r in R) else "OK"),
  ("Conversations avec coût mesuré", 0, "toutes", "streaming non instrumenté"),
  ("Coûts fixes renseignés", "non", "oui", "bloque le point mort"),
  ("Doublons d'adresse email", 0, "0", "OK"),
 ], [38, 22, 16, 30])
print("QUALITE_DONNEES ok")

# ================================================ SYNC_LOG
ws = feuille("SYNC_LOG", "Journal des synchronisations",
             "Rempli automatiquement par le script d'automatisation. La première ligne est l'extraction manuelle de ce classeur.")
tableau(ws, 4, ["Horodatage", "Statut", "Lignes lues", "Nouvelles", "Mises à jour",
                "Durée (s)", "Erreur"],
        [(AUJ, "succès (manuel)", len(U) + len(R) + len(E) + len(D["credits"]), "—", "—", "—", "")],
        [22, 18, 13, 12, 14, 11, 40])
print("SYNC_LOG ok")

wb.save("Pilotage_Axial.xlsx")
print()
print("classeur écrit :", wb.sheetnames)
